"""Dump a sheet from an .xlsb or .xlsx to TSV/JSON on stdout.
usage: dump_sheet.py <file> <sheet-substring> [maxrows] [--json]
"""
import sys, json, os

path = sys.argv[1]
sheet_filter = sys.argv[2]
maxrows = int(sys.argv[3]) if len(sys.argv) > 3 and not sys.argv[3].startswith('--') else 10**9
as_json = '--json' in sys.argv

def rows_xlsb(path, sheet_filter, maxrows):
    from pyxlsb import open_workbook
    with open_workbook(path) as wb:
        names = [n for n in wb.sheets if sheet_filter.lower() in n.lower()]
        if not names:
            raise SystemExit(f"no sheet matching {sheet_filter!r} in {wb.sheets}")
        name = names[0]
        out = []
        with wb.get_sheet(name) as sheet:
            for i, row in enumerate(sheet.rows()):
                if i >= maxrows:
                    break
                vals = [c.v for c in row]
                out.append(vals)
        return name, out

def rows_xlsx(path, sheet_filter, maxrows):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = [n for n in wb.sheetnames if sheet_filter.lower() in n.lower()]
    if not names:
        raise SystemExit(f"no sheet matching {sheet_filter!r} in {wb.sheetnames}")
    name = names[0]
    ws = wb[name]
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= maxrows:
            break
        out.append(list(row))
    return name, out

name, rows = (rows_xlsb if path.lower().endswith('.xlsb') else rows_xlsx)(path, sheet_filter, maxrows)
sys.stderr.write(f"sheet={name} rows={len(rows)}\n")
if as_json:
    print(json.dumps(rows, default=str, indent=1))
else:
    for r in rows:
        while r and r[-1] is None:
            r.pop()
        print("\t".join("" if v is None else str(v).replace("\t", " ").replace("\n", "\\n") for v in r))
