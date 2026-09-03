"""
Build the R5 report: for every deprecated column bound by a component the R5 class
names, read the LIVE row and say whether it is removable now - or whether an author
has to unfilter or hide it first.

Why this is a SWEEP and not a restatement of the report
-------------------------------------------------------
Two independent problems make "just list the R5 rows" the wrong shape:

1. **R5's own definition excludes half the question.** The class is "Deprecated
   column bound *but not rendered*", so a deprecated column that IS shown was never
   enumerated. Asking "which components still show a deprecated column" therefore
   cannot be answered from the R5 rows - by construction the answer would be zero.

2. **R5's "hidden and not filtered" claim is not reliable.** Applying the class
   found components carrying ACTIVE FILTERS on the very column the report calls
   unfiltered. The likely cause is a config-shape blind spot: a v2 component keeps
   filters in a top-level `filters` key, a v1 component has no such key at all and
   uses `dataRequest.filterGroups`, so a check written against `filters` is
   vacuously true for every v1 component - the same mistake that hid half the
   pattern from the T6 fetch-mode scan.

So this script takes the SECTIONS R5 names, and sweeps each one for **every**
deprecated column it binds, re-deriving rendered/filtered/referenced from the live
payload. R5 rows are then joined onto that sweep by column name. A row the sweep
finds and R5 never listed is reported as such; an R5 row whose quoted title matches
no live source column is reported as a report defect, with the sweep's unclaimed
candidate named but never silently substituted.

    python build_r5_report.py <report>.csv --out-dir src/themes/mny/design/reports
                              [--basename county-template-qa-r5-deprecated-columns]
                              [--class R5] [--limit N]

Writes `<basename>.{csv,xlsx,html}`. Carries `Status` / `Assigned to` / `Date fixed`
/ `Notes` forward from a previous build of the same basename, so an owner's triage
survives a rebuild. Read-only - it never writes to the site.
"""
import argparse, collections, csv, datetime as dt, html, json, os, re, subprocess, sys

HERE = os.path.dirname(os.path.abspath(__file__))
DMS = os.path.abspath(os.path.join(HERE, '../../../../../src/dms/packages/dms/cli/bin/dms.js'))

TITLE_RE = re.compile(r'live title is "(.+?)"')

# How a source steward marks a column as retired, in the titles they actually use:
#   "(Delete) FEMA - Financial", "(Deprecated) Dam Rehabilitation/Removal",
#   "# Not Started - deprecated", "Hazards - (no flood, deprecated)",
#   "Hazards (Deprecated)", "Hazard Multi - Deprecate"
DEPRECATED_RE = re.compile(r'\(\s*delete\s*\)|deprecat', re.I)

# The binding source-schema snapshot: v2 under `externalSource`, v1 under
# `sourceInfo` (migrateToV2.js:203-226). A component that is itself consumable as a
# source caches the whole schema a THIRD time under outputSourceInfo.asUdaConfig.
BINDING_SNAPSHOT_KEYS = ('externalSource', 'sourceInfo')
EXTRA_SNAPSHOT_PATHS = (('outputSourceInfo', 'asUdaConfig', 'sourceInfo', 'columns'),)


def get_in(o, keys):
    for k in keys:
        if not isinstance(o, dict):
            return None
        o = o.get(k)
    return o


def binding_snapshot(edo):
    found = [k for k in BINDING_SNAPSHOT_KEYS
             if isinstance(edo.get(k), dict) and isinstance(edo[k].get('columns'), list)]
    return (found[0], edo[found[0]]['columns']) if len(found) == 1 else (None, None)


def snapshot_prefixes(edo, binding_key):
    """Paths holding a copy of the SOURCE schema rather than a binding.

    Extras are accepted only when their column-name set matches the binding
    snapshot's - never on the strength of the path looking snapshot-ish.
    """
    out = [f'{binding_key}.columns']
    want = {c.get('name') for c in edo[binding_key]['columns'] if isinstance(c, dict)}
    for p in EXTRA_SNAPSHOT_PATHS:
        cols = get_in(edo, list(p))
        if isinstance(cols, list) and {c.get('name') for c in cols if isinstance(c, dict)} == want:
            out.append('.'.join(p))
    return out


def walk_strings(node, prefix=''):
    if isinstance(node, str):
        yield prefix, node
    elif isinstance(node, list):
        for i, v in enumerate(node):
            yield from walk_strings(v, f'{prefix}[{i}]')
    elif isinstance(node, dict):
        for k, v in node.items():
            yield (f'{prefix}{{key}}' if prefix else '{key}'), k
            yield from walk_strings(v, f'{prefix}.{k}' if prefix else k)


def find_filter_entries(node, col_name, prefix=''):
    """Every config object whose `col` is this column, wherever it lives.

    Structural on purpose: v2 keeps filters in a top-level `filters` tree, v1 in
    `dataRequest.filterGroups` / `lastDataRequest.filterGroups`. Matching on a key
    name is exactly what made the report's "not filtered" claim vacuous for v1.
    """
    out = []
    if isinstance(node, dict):
        if node.get('col') == col_name:
            out.append((prefix, node))
        for k, v in node.items():
            out.extend(find_filter_entries(v, col_name, f'{prefix}.{k}' if prefix else k))
    elif isinstance(node, list):
        for i, v in enumerate(node):
            out.extend(find_filter_entries(v, col_name, f'{prefix}[{i}]'))
    return out


def categorise(path):
    if path.startswith(('data[', 'data.')):
        return 'cached data rows'
    if path.startswith('display.'):
        return 'display config'
    if path.startswith('outputSourceInfo.columns'):
        return 'output schema'
    if path.startswith('join.'):
        return 'join'
    if path.startswith('columns['):
        return 'another bound column'
    return 'other'


def section_element_data(section_id, cache):
    if section_id in cache:
        return cache[section_id]
    out = subprocess.run(['node', DMS, 'section', 'dump', str(section_id)],
                         capture_output=True, text=True, encoding='utf-8')
    if out.returncode != 0:
        raise SystemExit(f'dms section dump {section_id} failed: {out.stderr[-400:]}')
    s = out.stdout
    row = json.loads(s[s.index('{"id"'):])
    raw = row['data']['element']['element-data']
    cache[section_id] = (row, json.loads(raw) if isinstance(raw, str) else None,
                         raw if isinstance(raw, str) else '')
    return cache[section_id]


# ---------------------------------------------------------------- dispositions

DISPOSITIONS = [
    ('Ready - hidden & unfiltered', 'bg-emerald-50 text-emerald-900 border-emerald-300'),
    ('Blocked - rendered & filtered', 'bg-rose-50 text-rose-900 border-rose-300'),
    ('Blocked - rendered', 'bg-orange-50 text-orange-900 border-orange-300'),
    ('Blocked - filtered', 'bg-mny-y50 text-mny-900 border-mny-y700'),
    ('Blocked - other reference', 'bg-purple-50 text-purple-900 border-purple-300'),
    ('Done - already unbound', 'bg-mny-100 text-mny-700 border-mny-200'),
    ('Report defect - title not in source', 'bg-slate-100 text-slate-700 border-slate-300'),
]
DISPO_CLASS = dict(DISPOSITIONS)
ACTION = {
    'Ready - hidden & unfiltered': 'None - remove it',
    'Blocked - rendered & filtered': 'Remove the filter, then hide the column',
    'Blocked - rendered': 'Hide the column (eye icon)',
    'Blocked - filtered': 'Remove the filter on this column',
    'Blocked - other reference': 'Investigate the reference first',
    'Done - already unbound': 'None - done',
    'Report defect - title not in source': 'None - fix the report',
}
MEANING = {
    'Ready - hidden & unfiltered': 'Remove now - no visible change to the page.',
    'Blocked - rendered & filtered': 'Shown AND filtered. Remove the filter, then hide it.',
    'Blocked - rendered': 'The column is on the page - hiding it is a content decision.',
    'Blocked - filtered': 'An active filter restricts the data to this column.',
    'Blocked - other reference': 'The name is wired into other config - investigate.',
    'Done - already unbound': 'Nothing to do; the column is not bound.',
    'Report defect - title not in source': 'The report quotes a title the source does not carry.',
}


def assess(edo, raw, binding_key, idx, entry):
    """rendered / filtered / other-reference facts for one bound column."""
    name = entry.get('name')
    rendered = entry.get('show') is True

    fentries = find_filter_entries(edo, name)
    active = [(p, o) for p, o in fentries if o.get('value') not in (None, [], '', {})]
    empty = [(p, o) for p, o in fentries if (p, o) not in active]

    snap_pre = tuple(snapshot_prefixes(edo, binding_key))
    own_pre = f'columns[{idx}]'
    fpaths = {p for p, _ in fentries}
    others = collections.OrderedDict()
    for path, value in walk_strings(edo):
        if name not in value:
            continue
        if path.startswith(snap_pre) or path.startswith(own_pre):
            continue
        if any(path.startswith(fp) for fp in fpaths):
            continue
        others.setdefault(categorise(path), []).append(path)

    # cached rows exist only for a fetched (i.e. shown) column - a consequence of
    # `show`, not an independent blocker
    blocking = {k: v for k, v in others.items() if k != 'cached data rows'}

    if blocking:
        d = 'Blocked - other reference'
        why = ('the column name is wired into '
               + ', '.join(f'{k} ({v[0]})' for k, v in blocking.items())
               + ' - removing it would leave that reference dangling')
    elif rendered and active:
        d, why = ('Blocked - rendered & filtered',
                  'the column is DISPLAYED and an active filter restricts the data to it')
    elif active:
        d, why = ('Blocked - filtered',
                  'the column is hidden, but an active filter restricts this component\'s data to '
                  'it - removing the binding would leave the query filtering on a column the '
                  'component no longer declares')
    elif rendered:
        d, why = ('Blocked - rendered',
                  'show=true - the column is on the page, so removing it changes what readers see')
    else:
        d, why = ('Ready - hidden & unfiltered',
                  'not shown, no active filter, no other reference in the payload - removable '
                  'with no visible change')

    delta = ''
    if d == 'Ready - hidden & unfiltered' and json.dumps(json.loads(raw)) == raw:
        nxt = json.loads(raw)
        nxt['columns'].pop(idx)
        delta = len(json.dumps(nxt)) - len(raw)

    return {
        'show': 'true' if rendered else ('false' if entry.get('show') is False else '(unset)'),
        'Filtered': 'Yes' if active else ('empty filter' if empty else 'No'),
        'Filter entries': len(fentries),
        'Filter detail': ' | '.join(f'{p} value={json.dumps(o.get("value"))}'
                                    for p, o in (active or empty))[:600],
        'Other reference count': sum(len(v) for v in blocking.values()),
        'Other references': ' | '.join(f'{k}: {len(v)} ({v[0]})' for k, v in others.items())[:600],
        'Disposition': d, 'Why': why, 'Author action': ACTION[d], 'Payload delta': delta,
    }


FIELDS = ['Fix ID', 'Likely Fix ID', 'In R5 report', 'Disposition', 'Author action', 'Why', 'Page URL', 'Page',
          'Page ID', 'Section title', 'Draft section ID', 'Component kind', 'Config shape',
          'Source snapshot key', 'Source', 'Live column title', 'Column name', 'Stored title',
          'Title stale', 'Bound index', 'Bound columns', 'show', 'Filtered', 'Filter entries',
          'Filter detail', 'Other reference count', 'Other references', 'Payload delta',
          'Status', 'Assigned to', 'Date fixed', 'Notes']
CARRY = ('Status', 'Assigned to', 'Date fixed', 'Notes')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('report')
    ap.add_argument('--out-dir', required=True)
    ap.add_argument('--basename', default='county-template-qa-r5-deprecated-columns')
    ap.add_argument('--klass', '--class', dest='klass', default='R5')
    ap.add_argument('--limit', type=int)
    args = ap.parse_args()

    with open(args.report, encoding='utf-8-sig', newline='') as f:
        src_rows = list(csv.DictReader(f))
    r5 = [r for r in src_rows if (r.get('Class') or r.get('Type')) == args.klass]
    if args.limit:
        r5 = r5[:args.limit]
    if not r5:
        raise SystemExit(f'no rows with Class == {args.klass!r} in {args.report}')

    prev = {}
    prev_path = os.path.join(args.out_dir, args.basename + '.csv')
    if os.path.exists(prev_path):
        with open(prev_path, encoding='utf-8-sig', newline='') as f:
            for r in csv.DictReader(f):
                key = (r['Draft section ID'], r.get('Column name', ''))
                prev[key] = {k: r.get(k, '') for k in CARRY}

    by_section = collections.OrderedDict()
    for r in r5:
        by_section.setdefault(r['Draft section ID'], []).append(r)

    cache, out_rows = {}, []
    for n, (sid, rows) in enumerate(by_section.items(), 1):
        meta = rows[0]
        row, edo, raw = section_element_data(sid, cache)
        ident = {
            'Page URL': meta.get('Page URL', ''), 'Page': meta.get('Page', ''),
            'Page ID': meta.get('Page ID', ''), 'Section title': meta.get('Section title', ''),
            'Draft section ID': sid, 'Component kind': meta.get('Component kind', ''),
            'Config shape': meta.get('Config shape', ''), 'Source': meta.get('Source', ''),
        }
        binding_key, snap_cols = binding_snapshot(edo) if edo else (None, None)
        ident['Source snapshot key'] = binding_key or '(none)'
        if binding_key is None:
            out_rows.append({**ident, 'Fix ID': ', '.join(r['Fix ID'] for r in rows),
                             'In R5 report': 'Yes',
                             'Disposition': 'Report defect - title not in source',
                             'Author action': ACTION['Report defect - title not in source'],
                             'Why': 'no single source snapshot with a columns list on this row'})
            continue

        title_of = {c.get('name'): c.get('display_name') for c in snap_cols if isinstance(c, dict)}

        # ---- what the R5 rows claim, resolved against the live source schema
        claimed, unresolved = {}, []
        for r in rows:
            m = TITLE_RE.search(r.get('What is wrong', ''))
            t = m.group(1) if m else ''
            hits = [c for c in snap_cols if c.get('display_name') == t]
            if len(hits) == 1:
                claimed[hits[0]['name']] = (r['Fix ID'], t)
            else:
                unresolved.append((r['Fix ID'], t, len(hits)))

        # ---- the SWEEP: every deprecated column this component actually binds
        bound = [(i, c) for i, c in enumerate(edo.get('columns', [])) if isinstance(c, dict)]
        swept = [(i, c) for i, c in bound
                 if DEPRECATED_RE.search(title_of.get(c.get('name')) or '')]

        # A swept column no R5 row resolved to, and an R5 row whose title resolved
        # to nothing, are usually the SAME finding seen from both sides - the report
        # mangled the title. Pair them only when each side has exactly one
        # candidate, and label the pairing as inferred; never merge them silently.
        unclaimed = [n for n in (c.get('name') for _, c in swept) if n not in claimed]
        likely = {}
        if len(unresolved) == 1 and len(unclaimed) == 1:
            likely[unclaimed[0]] = unresolved[0][0]

        for i, entry in swept:
            name = entry.get('name')
            live_title = title_of.get(name) or ''
            fix_id, _ = claimed.pop(name, ('', ''))
            if fix_id:
                in_r5 = 'Yes'
            elif name in likely:
                in_r5 = f'Title mismatch - likely {likely[name]}'
            else:
                in_r5 = 'No - found by sweep'
            rec = {
                **ident,
                'Fix ID': fix_id,
                'Likely Fix ID': likely.get(name, ''),
                'In R5 report': in_r5,
                'Live column title': live_title, 'Column name': name,
                'Stored title': entry.get('display_name', ''),
                'Title stale': 'Yes' if entry.get('display_name') != live_title else 'No',
                'Bound index': i, 'Bound columns': len(bound),
                **assess(edo, raw, binding_key, i, entry),
            }
            if name in likely:
                rec['Why'] += (f' Not listed under this title by R5: {likely[name]} names it '
                               f'"{unresolved[0][1]}", which no source column carries - see that '
                               f'row. The pairing is inferred from both sides having exactly one '
                               f'candidate, and is UNCONFIRMED.')
            out_rows.append(rec)

        # ---- R5 rows whose column resolved but is no longer bound
        for name, (fix_id, t) in claimed.items():
            out_rows.append({
                **ident, 'Fix ID': fix_id, 'In R5 report': 'Yes',
                'Live column title': t, 'Column name': name,
                'Bound columns': len(bound),
                'Disposition': 'Done - already unbound',
                'Author action': ACTION['Done - already unbound'],
                'Why': 'the column is not in this component\'s columns[] - already removed',
            })

        # ---- R5 rows whose quoted title matches no live source column
        for fix_id, t, nhits in unresolved:
            cand = ''
            if len(unclaimed) == 1:
                cand = (f' The sweep finds exactly one unclaimed deprecated column on this '
                        f'component - "{title_of.get(unclaimed[0])}" - which is a likely but '
                        f'UNCONFIRMED match; it is listed as its own row, and this row is a '
                        f'duplicate of it rather than a second finding.')
            out_rows.append({
                **ident, 'Fix ID': fix_id, 'In R5 report': 'Yes',
                'Live column title': t, 'Column name': '',
                'Bound columns': len(bound),
                'Disposition': 'Report defect - title not in source',
                'Author action': ACTION['Report defect - title not in source'],
                'Why': (f'{nhits} columns in {binding_key}.columns are titled "{t}" - expected '
                        f'exactly 1. The report quotes a title the source does not carry.' + cand),
            })

        print(f'[{n}/{len(by_section)}] {sid}  {len(swept)} deprecated column(s) bound, '
              f'{len(rows)} R5 row(s)', file=sys.stderr)

    for rec in out_rows:
        for k, v in (prev.get((rec['Draft section ID'], rec.get('Column name', ''))) or {}).items():
            if v:
                rec[k] = v

    os.makedirs(args.out_dir, exist_ok=True)
    base = os.path.join(args.out_dir, args.basename)
    write_csv(base + '.csv', out_rows)
    write_xlsx(base + '.xlsx', out_rows)
    write_html(base + '.html', out_rows, args.report, len(r5))

    tally = collections.Counter(r['Disposition'] for r in out_rows)
    print(f'\n{len(out_rows)} rows over {len(by_section)} sections '
          f'(from {len(r5)} R5 report rows)')
    for d, _ in DISPOSITIONS:
        if tally.get(d):
            print(f'  {tally[d]:>3}  {d}')
    mismatch = sum(1 for r in out_rows if r.get('Likely Fix ID'))
    unlisted = sum(1 for r in out_rows if r.get('In R5 report', '').startswith('No'))
    print(f'\n  {mismatch} column(s) R5 lists under a title the source does not carry '
          f'(paired back to a Fix ID, unconfirmed)')
    print(f'  {unlisted} column(s) R5 does not list at all')
    print(f'\n-> {base}.csv / .xlsx / .html')


def write_csv(path, rows):
    with open(path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=FIELDS, extrasaction='ignore')
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, '') for k in FIELDS})


FILLS = {
    'Ready - hidden & unfiltered': 'DFF5E7',
    'Blocked - rendered & filtered': 'F8D9D8',
    'Blocked - rendered': 'FBE3D2',
    'Blocked - filtered': 'FCF0D8',
    'Blocked - other reference': 'EADFF5',
    'Done - already unbound': 'E8EEF1',
    'Report defect - title not in source': 'ECECEC',
}


def write_xlsx(path, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hdr_fill = PatternFill('solid', fgColor='2D3E4C')
    thin = Side(style='thin', color='C5D7E0')

    def header(ws, cols, widths):
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.fill = hdr_fill
            c.alignment = Alignment(vertical='center', wrap_text=True)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    ws = wb.active
    ws.title = 'Deprecated columns'
    widths = {'Fix ID': 10, 'Likely Fix ID': 13, 'In R5 report': 26, 'Disposition': 30, 'Author action': 32, 'Why': 70,
              'Page URL': 14, 'Page': 34, 'Page ID': 10, 'Section title': 30,
              'Draft section ID': 15, 'Component kind': 14, 'Config shape': 12,
              'Source snapshot key': 17, 'Source': 22, 'Live column title': 34, 'Column name': 44,
              'Stored title': 26, 'Title stale': 11, 'Bound index': 11, 'Bound columns': 13,
              'show': 9, 'Filtered': 12, 'Filter entries': 12, 'Filter detail': 50,
              'Other reference count': 12, 'Other references': 44, 'Payload delta': 12,
              'Status': 12, 'Assigned to': 14, 'Date fixed': 12, 'Notes': 60}
    header(ws, FIELDS, [widths.get(k, 16) for k in FIELDS])
    for r in rows:
        ws.append([r.get(k, '') for k in FIELDS])
        fill = FILLS.get(r['Disposition'])
        for c in ws[ws.max_row]:
            if fill:
                c.fill = PatternFill('solid', fgColor=fill)
            c.border = Border(bottom=thin)
            c.alignment = Alignment(vertical='top',
                                    wrap_text=c.column_letter in ('C', 'D', 'E', 'AE'))
    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(FIELDS))}{ws.max_row}'

    ws2 = wb.create_sheet('Summary')
    tally = collections.Counter(r['Disposition'] for r in rows)
    header(ws2, ['Disposition', 'Columns', 'Sections', 'What it means'], (36, 10, 10, 66))
    for d, _ in DISPOSITIONS:
        if not tally.get(d):
            continue
        secs = len({r['Draft section ID'] for r in rows if r['Disposition'] == d})
        ws2.append([d, tally[d], secs, MEANING[d]])
        ws2.cell(ws2.max_row, 1).fill = PatternFill('solid', fgColor=FILLS[d])
    ws2.append([])
    ws2.append(['TOTAL', len(rows), len({r['Draft section ID'] for r in rows}), ''])
    ws2.cell(ws2.max_row, 1).font = Font(bold=True)

    ws3 = wb.create_sheet('Author worklist')
    header(ws3, ['Page', 'Section title', 'Draft section ID', 'Deprecated column (live title)',
                 'Column name', 'What to do', 'Detail', 'Fix ID'],
           (34, 30, 15, 36, 44, 34, 56, 10))
    work = [r for r in rows if r['Disposition'].startswith('Blocked')]
    for r in sorted(work, key=lambda x: (x['Page'], x['Section title'], x.get('Column name', ''))):
        ws3.append([r['Page'], r['Section title'], r['Draft section ID'],
                    r.get('Live column title', ''), r.get('Column name', ''),
                    r.get('Author action', ''),
                    r.get('Filter detail', '') or r.get('Other references', ''), r['Fix ID']])
        ws3.cell(ws3.max_row, 6).fill = PatternFill('solid', fgColor=FILLS[r['Disposition']])
    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f'A1:H{ws3.max_row}'

    ws4 = wb.create_sheet('Ready to remove')
    header(ws4, ['Fix ID', 'Page', 'Section title', 'Draft section ID', 'Column name',
                 'Live column title', 'Expected payload delta'],
           (10, 34, 30, 15, 44, 36, 20))
    for r in [x for x in rows if x['Disposition'] == 'Ready - hidden & unfiltered']:
        ws4.append([r['Fix ID'], r['Page'], r['Section title'], r['Draft section ID'],
                    r.get('Column name', ''), r.get('Live column title', ''),
                    r.get('Payload delta', '')])
    ws4.freeze_panes = 'A2'
    ws4.auto_filter.ref = f'A1:G{ws4.max_row}'

    wb.save(path)


def write_html(path, rows, report_src, n_r5):
    e = html.escape
    tally = collections.Counter(r['Disposition'] for r in rows)
    secs = len({r['Draft section ID'] for r in rows})
    ready = tally.get('Ready - hidden & unfiltered', 0)
    blocked = sum(v for k, v in tally.items() if k.startswith('Blocked'))
    shown = tally.get('Blocked - rendered', 0) + tally.get('Blocked - rendered & filtered', 0)
    filt = tally.get('Blocked - filtered', 0) + tally.get('Blocked - rendered & filtered', 0)
    unlisted = sum(1 for r in rows if r.get('In R5 report', '').startswith('No'))
    mismatch = sum(1 for r in rows if r.get('Likely Fix ID'))

    def pill(text, cls):
        return (f'<span class="inline-block whitespace-nowrap px-2 py-[1px] rounded border '
                f'text-[11px] font-semibold tracking-wide {cls}">{e(text)}</span>')

    trs = []
    for r in rows:
        d = r['Disposition']
        nm = r.get('Column name') or ''
        extra = ''
        if r.get('Filter detail') and r.get('Filtered') == 'Yes':
            extra += f'<div class="text-mny-400 mt-1"><code>{e(r["Filter detail"])}</code></div>'
        if r.get('Other reference count'):
            extra += f'<div class="text-mny-400 mt-1"><code>{e(r.get("Other references",""))}</code></div>'
        trs.append(
            f'<tr class="border-t border-mny-100" data-d="{e(d)}" data-r5="{e(r.get("In R5 report",""))}">'
            f'<td class="py-1.5 pr-3 pl-4 align-top"><code>{e(r["Fix ID"]) or "&mdash;"}</code>'
            + ('' if r.get('In R5 report') == 'Yes'
               else '<div class="text-[10px] font-semibold text-purple-700 uppercase tracking-wide">sweep</div>')
            + f'</td>'
            f'<td class="py-1.5 pr-3 align-top">{pill(d, DISPO_CLASS[d])}</td>'
            f'<td class="py-1.5 pr-3 align-top text-[12.5px]">{e(r.get("Author action",""))}</td>'
            f'<td class="py-1.5 pr-3 align-top text-[12.5px]"><div class="font-semibold">'
            f'{e(r.get("Section title") or "(untitled)")}</div>'
            f'<div class="text-mny-400"><code>{e(r["Page"])}</code> &middot; '
            f'<code>{e(r["Draft section ID"])}</code></div></td>'
            f'<td class="py-1.5 pr-3 align-top text-[12.5px]">{e(r.get("Live column title",""))}'
            f'<div class="text-mny-400"><code>{e(nm[:64])}{"&hellip;" if len(nm)>64 else ""}</code></div></td>'
            f'<td class="py-1.5 pr-3 align-top num text-[12.5px]">{e(str(r.get("show","")))}</td>'
            f'<td class="py-1.5 pr-3 align-top num text-[12.5px]">{e(str(r.get("Filtered","")))}</td>'
            f'<td class="py-1.5 pr-4 align-top text-[12px] text-mny-700">{e(r.get("Why",""))}{extra}</td>'
            f'</tr>')

    summary = ''.join(
        f'<tr class="border-t border-mny-100"><td class="py-1.5 pr-3">{pill(d, DISPO_CLASS[d])}</td>'
        f'<td class="py-1.5 pr-3 num font-semibold">{tally[d]}</td>'
        f'<td class="py-1.5 pr-3 num">{len({r["Draft section ID"] for r in rows if r["Disposition"]==d})}</td>'
        f'<td class="py-1.5 text-[13px] text-mny-700">{e(MEANING[d])}</td></tr>'
        for d, _ in DISPOSITIONS if tally.get(d))

    btns = ''.join(
        f'<button data-f="{e(d)}" class="fbtn px-2.5 py-1 rounded border text-[12px] '
        f'font-semibold {DISPO_CLASS[d]}">{e(d)} <span class="num">{tally[d]}</span></button>'
        for d, _ in DISPOSITIONS if tally.get(d))

    doc = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>MitigateNY &mdash; R5 Deprecated Columns</title>
<script src="https://cdn.tailwindcss.com"></script>
<script>
tailwind.config = {{ theme: {{ extend: {{
  colors: {{'mny-900':'#2D3E4C','mny-700':'#37576B','mny-400':'#6D96AE','mny-200':'#C5D7E0',
           'mny-100':'#E0EBF0','mny-50':'#F3F8F9','mny-y700':'#EAAD43','mny-y500':'#F1CA87',
           'mny-y50':'#FCF6EC','mny-red':'#DD524C','mny-redk':'#AA2E26','mny-grn':'#54B99B'}},
  fontFamily: {{display:['"Oswald"','sans-serif'], proxima:['"Source Sans 3"','system-ui','sans-serif']}}
}}}}}}
</script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Oswald:wght@400;500;600;700&family=Source+Sans+3:ital,wght@0,400;0,600;0,700;1,400&display=swap">
<style>
  body{{font-family:'Source Sans 3',system-ui,sans-serif}}
  h1,h2,h3,.display{{font-family:'Oswald',sans-serif;letter-spacing:.01em}}
  code{{font-family:ui-monospace,'Cascadia Mono',Consolas,monospace;font-size:.86em;overflow-wrap:anywhere}}
  .wrap{{max-width:1600px}}
  table{{border-collapse:separate;border-spacing:0}}
  thead th{{position:sticky;top:0;z-index:2;background:#fff}}
  tbody tr:hover{{background:#F3F8F9}}
  .num{{font-variant-numeric:tabular-nums}}
  .fbtn.off{{opacity:.32}}
  @media print{{.noprint{{display:none}}thead th{{position:static}}}}
</style>
</head>
<body class="bg-[#F4F4F4] text-mny-900">
<div class="wrap mx-auto px-5 py-8">

<header class="mb-8 border-b-2 border-mny-700 pb-5">
  <div class="text-[11px] font-semibold tracking-[.18em] text-mny-400 uppercase">MitigateNY &middot; County template QA</div>
  <h1 class="text-4xl font-semibold mt-1">R5 &mdash; Deprecated columns, readiness to remove</h1>
  <p class="text-mny-700 mt-2 max-w-4xl">Every deprecated column bound by the
  <strong>{secs}</strong> components the <strong>R5</strong> class names in
  <code>{e(os.path.basename(report_src))}</code> &mdash; re-derived from the <strong>live</strong>
  row. Is the column hidden and unfiltered, and therefore removable with no visible change &mdash;
  or does an author have to unfilter or hide it first? Built {dt.date.today():%Y-%m-%d}.
  Read-only; nothing here writes to the site.</p>
</header>

<section class="mb-8 bg-white border-2 border-mny-red rounded-lg p-5">
  <div class="text-[11px] font-semibold tracking-[.18em] text-mny-red uppercase">Why this is a sweep, not a restatement</div>
  <h2 class="text-2xl font-semibold mt-1 mb-3">The R5 rows cannot answer this question on their own</h2>
  <div class="grid md:grid-cols-2 gap-5 text-[13.5px] text-mny-700">
    <div>
      <p><strong>1. R5&rsquo;s definition excludes half the question.</strong> The class is
      &ldquo;Deprecated column bound <em>but not rendered</em>&rdquo;, so a deprecated column that
      <em>is</em> shown was never enumerated. Asking &ldquo;which components still show a deprecated
      column&rdquo; against the R5 rows would return zero <em>by construction</em>, not by fact.</p>
      <p class="mt-2">So this report starts from the <strong>sections</strong> R5 names and sweeps
      each for <em>every</em> deprecated column it binds &mdash; a column whose live source title
      carries <code>(Delete)</code> or <code>deprecated</code>. R5 rows are joined onto that sweep
      by column name &mdash; <strong>not</strong> by the title text.</p>
      <p class="mt-2">Of the deprecated columns the sweep found, <strong>{mismatch}</strong>
      are ones R5 <em>does</em> list, but under a title the source does not carry &mdash; the
      name join is what recovers them. Each is paired back to its <code>Likely Fix ID</code>
      and marked <strong>unconfirmed</strong>, never merged silently.
      <strong>{unlisted}</strong> are columns R5 does not list at all.</p>

    </div>
    <div>
      <p><strong>2. R5&rsquo;s &ldquo;hidden and not filtered&rdquo; claim is not reliable.</strong>
      Applying the class found components carrying <strong>active filters</strong> on the very
      column the report calls unfiltered. The likely cause is a config-shape blind spot: a
      <strong>v2</strong> component keeps filters in a top-level <code>filters</code> key, a
      <strong>v1</strong> component has <em>no such key at all</em> and uses
      <code>dataRequest.filterGroups</code> &mdash; so a check written against <code>filters</code>
      is <strong>vacuously true for every v1 component</strong>. The same mistake that hid half the
      pattern from the T6 fetch-mode scan.</p>
      <p class="mt-2">Here filters are found <strong>structurally</strong> &mdash; any config object
      whose <code>col</code> is this column, wherever it lives &mdash; and an empty
      <code>value</code> is reported as an inactive filter rather than an active one.</p>
    </div>
  </div>
  <p class="text-[13px] text-mny-700 mt-4 max-w-5xl">Every other mention of the column name in the
  payload is reported too, because a splice out of <code>columns[]</code> is index-safe but
  <strong>not reference-safe</strong>. Source-schema snapshots
  (<code>externalSource.columns</code>, <code>sourceInfo.columns</code>, and
  <code>outputSourceInfo.asUdaConfig.sourceInfo.columns</code> on components that are themselves
  consumable as a source) are excluded &mdash; they describe the source, not the binding, and they
  legitimately list the deprecated column.</p>
</section>

<section class="mb-8 grid md:grid-cols-4 gap-4">
  <div class="bg-white border border-mny-200 rounded-lg p-5">
    <div class="text-[11px] font-semibold tracking-[.18em] text-mny-400 uppercase">Removable now</div>
    <div class="text-5xl font-semibold mt-1 text-emerald-700 num">{ready}</div>
    <p class="text-[13px] text-mny-700 mt-2">hidden, unfiltered, referenced nowhere else. Straight
    through the fix loop.</p>
  </div>
  <div class="bg-white border border-mny-200 rounded-lg p-5">
    <div class="text-[11px] font-semibold tracking-[.18em] text-mny-400 uppercase">Still filtering</div>
    <div class="text-5xl font-semibold mt-1 text-mny-y700 num">{filt}</div>
    <p class="text-[13px] text-mny-700 mt-2">an active filter restricts the component&rsquo;s data
    to the deprecated column. Unfilter first.</p>
  </div>
  <div class="bg-white border border-mny-200 rounded-lg p-5">
    <div class="text-[11px] font-semibold tracking-[.18em] text-mny-400 uppercase">Still showing</div>
    <div class="text-5xl font-semibold mt-1 text-orange-700 num">{shown}</div>
    <p class="text-[13px] text-mny-700 mt-2"><code>show=true</code> &mdash; the column is on the
    page. Hiding it is a content decision.</p>
  </div>
  <div class="bg-white border border-mny-200 rounded-lg p-5">
    <div class="text-[11px] font-semibold tracking-[.18em] text-mny-400 uppercase">Scope</div>
    <div class="text-5xl font-semibold mt-1 num">{len(rows)}</div>
    <p class="text-[13px] text-mny-700 mt-2">bound deprecated columns over <strong>{secs}</strong>
    sections, from <strong>{n_r5}</strong> R5 rows. One row = one column on one component.</p>
  </div>
</section>

<section class="mb-8 bg-white border border-mny-200 rounded-lg p-5">
  <h2 class="text-lg font-semibold mb-2">Disposition &mdash; {blocked} of {len(rows)} need an authoring pass first</h2>
  <table class="text-sm w-full">
    <thead class="text-left text-mny-400 text-[11px] uppercase tracking-wider">
      <tr><th class="py-1 pr-3">Verdict</th><th class="py-1 pr-3 num">Columns</th>
          <th class="py-1 pr-3 num">Sections</th><th class="py-1">What it means</th></tr></thead>
    <tbody>{summary}</tbody>
  </table>
  <p class="text-[13px] text-mny-700 mt-3">The <code>.xlsx</code> carries this as three sheets:
  <strong>Deprecated columns</strong> (everything), <strong>Author worklist</strong> (only the rows
  needing a hide or an unfilter, sorted by page), and <strong>Ready to remove</strong> (the fix-loop
  input, with each row&rsquo;s expected payload delta as a write signature).</p>
</section>

<section class="mb-4 noprint flex flex-wrap gap-2 items-center">
  <span class="text-[11px] font-semibold tracking-[.18em] text-mny-400 uppercase mr-1">Filter</span>
  {btns}
  <button id="all" class="px-2.5 py-1 rounded border border-mny-700 bg-mny-700 text-white text-[12px] font-semibold">Show all</button>
  <input id="q" type="search" placeholder="search page, column, section&hellip;"
         class="ml-auto px-3 py-1 rounded border border-mny-200 text-[13px] w-72">
</section>

<section class="bg-white border border-mny-200 rounded-lg overflow-hidden">
  <div class="overflow-x-auto">
  <table class="text-sm w-full">
    <thead class="text-left text-mny-400 text-[11px] uppercase tracking-wider border-b border-mny-200">
      <tr><th class="py-2 pr-3 pl-4">Fix</th><th class="py-2 pr-3">Verdict</th>
          <th class="py-2 pr-3">Author action</th><th class="py-2 pr-3">Component</th>
          <th class="py-2 pr-3">Deprecated column</th><th class="py-2 pr-3">show</th>
          <th class="py-2 pr-3">filtered</th><th class="py-2 pr-4">Why</th></tr>
    </thead>
    <tbody id="rows">{''.join(trs)}</tbody>
  </table>
  </div>
</section>

<footer class="mt-8 text-[12px] text-mny-400">
  Generated by <code>planning/mitigateny/skills/scripts/report_fixes/build_r5_report.py</code>
  from <code>{e(os.path.basename(report_src))}</code>, reading live rows through the DMS CLI.
  The fix loop is <code>planning/mitigateny/skills/applying-report-fixes-to-a-live-site.md</code> &sect;2e.
</footer>

</div>
<script>
const rows=[...document.querySelectorAll('#rows tr')];
let active=new Set();
function apply(){{
  const q=document.getElementById('q').value.toLowerCase();
  rows.forEach(r=>{{
    const okD=!active.size||active.has(r.dataset.d);
    const okQ=!q||r.textContent.toLowerCase().includes(q);
    r.style.display=(okD&&okQ)?'':'none';
  }});
  document.querySelectorAll('.fbtn').forEach(b=>
    b.classList.toggle('off', active.size>0 && !active.has(b.dataset.f)));
}}
document.querySelectorAll('.fbtn').forEach(b=>b.onclick=()=>{{
  active.has(b.dataset.f)?active.delete(b.dataset.f):active.add(b.dataset.f); apply();
}});
document.getElementById('all').onclick=()=>{{active.clear();document.getElementById('q').value='';apply();}};
document.getElementById('q').oninput=apply;
</script>
</body>
</html>
'''
    with open(path, 'w', encoding='utf-8', newline='\n') as f:
        f.write(doc)


if __name__ == '__main__':
    main()
