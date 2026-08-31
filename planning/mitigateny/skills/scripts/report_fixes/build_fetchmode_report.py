# -*- coding: utf-8 -*-
"""
Turn one or more `scan_fetchmode.mjs` outputs into the T6 fetch-mode tracking
report: a CSV, an XLSX with editable triage columns, and a self-contained HTML
report.

Constant-free: every value comes from the scan JSON or a flag. Pass several
`--scan` arguments to put `county_template` and its duplicates in one document
(the `Pattern ID` column keeps them apart, exactly as the T5 tag report does).

The recommended fix implements the owner's rule:

    external source  ->  Smart (fetch on change)
    internal source  ->  Force (always re-fetch)

SCOPE (narrowed 2026-08-28). `Force` re-queries on every mount, so every extra
component set to it is page-load cost. The owner narrowed the sweep to the
datasets an author actually edits and expects to see change, on the pages that
matter:

  --source-allow    only these source names are fixed. Everything else -
                    external sources, and internal ones like LHMP_IA that are
                    large and effectively static - is DEFERRED.
  --page-allow-from only pages a companion report labels in-scope are fixed.

Deferred rows are NOT dropped. They keep their Fix ID and every scanned fact
and move to a `Deferred` bucket carrying the reason, because "we chose not to
fix this" and "we never looked at it" must not look the same - the same
principle already applied to components with no control at all.

Scope: only Card / Spreadsheet / Graph carry a `Data Fetch Mode` control
(Card.config.jsx, spreadsheet/config.jsx, graph/config.jsx,
graph_new/config.jsx). Filter and Header bind a source but expose no such
setting, so they cannot be fixed by an author at all - they are emitted to a
separate out-of-scope section rather than dropped, because "not in the catalog"
and "does not exist" must not look the same.

usage:
  python build_fetchmode_report.py --scan <scan.json> [--scan <scan2.json> ...]
         --out-dir <reports dir> --slug county-template-qa-t6-fetchmode
         [--subdomain-map 1300890=county_template,2249247=suffolk_draft]
         [--parent-finding T6-001] [--date YYYY-MM-DD]
         [--source-allow Actions_Revised,Hazards_of_Concern,...]
         [--page-allow-from <t5.xlsx> --page-allow-sheet '<sheet>'
          --page-allow-column 'Page Relevance' --page-allow-value Relevant]
"""
import sys, os, json, csv, html, argparse, collections

IN_SCOPE = ('Card', 'Spreadsheet', 'Graph')
MODE_LABEL = {
    'cache': 'Cache (use preloaded data)',
    'smart': 'Smart (fetch on change)',
    'force': 'Force (always re-fetch)',
}
WANT = {'external': 'smart', 'internal': 'force'}

ap = argparse.ArgumentParser()
ap.add_argument('--scan', action='append', required=True)
ap.add_argument('--out-dir', required=True)
ap.add_argument('--slug', required=True)
ap.add_argument('--subdomain-map', default='')
ap.add_argument('--parent-finding', default='T6-001')
ap.add_argument('--date', default='')
ap.add_argument('--id-ledger', default='',
                help='JSON file mapping "<patternId>:<sectionId>" -> Fix ID. Read before assigning '
                     'and rewritten after, so a Fix ID always means the same component even as the '
                     'pattern gains and loses components. STRONGLY recommended - without it Fix IDs '
                     'are positional and renumber whenever anything upstream changes.')
ap.add_argument('--source-allow', default='',
                help='comma-separated data source NAMES that are in scope. Anything else is '
                     'deferred with a reason rather than dropped. Omit to keep every source in '
                     'scope (the pre-2026-08-28 behaviour).')
ap.add_argument('--page-allow-from', default='',
                help='CSV or XLSX whose rows label pages as in- or out-of-scope. Pages not '
                     'labelled with --page-allow-value are deferred. Omit to keep every page.')
ap.add_argument('--page-allow-sheet', default='',
                help='sheet name inside --page-allow-from when it is an XLSX (default: first)')
ap.add_argument('--page-allow-column', default='Page Relevance',
                help='column in --page-allow-from holding the label')
ap.add_argument('--page-allow-value', default='Relevant',
                help='the value in --page-allow-column that means "in scope"')
ap.add_argument('--page-key-column', default='Page',
                help='column in --page-allow-from holding the page slug')
ap.add_argument('--no-carry', action='store_true',
                help='do not carry forward Assigned to / Notes from an existing CSV, and accept a '
                     'Fix ID renumbering (use only when the renumbering is intended)')
args = ap.parse_args()

subdomains = {}
for pair in filter(None, args.subdomain_map.split(',')):
    k, _, v = pair.partition('=')
    subdomains[k.strip()] = v.strip()

# ── scope ─────────────────────────────────────────────────────────────────────
# Both filters are OPT-IN. With neither flag the report keeps its original
# meaning (every source, every page), so an old command line still reproduces
# the old document.
source_allow = set(x.strip() for x in args.source_allow.split(',') if x.strip())

page_allow = None
if args.page_allow_from:
    src, page_allow = args.page_allow_from, set()
    seen_pages = set()
    if src.lower().endswith(('.xlsx', '.xlsm')):
        import openpyxl
        wb_in = openpyxl.load_workbook(src, read_only=True, data_only=True)
        ws_in = wb_in[args.page_allow_sheet] if args.page_allow_sheet else wb_in.worksheets[0]
        it = ws_in.iter_rows(values_only=True)
        hdr = ['' if h is None else str(h) for h in next(it)]
        recs = [{hdr[i]: ('' if v is None else str(v)) for i, v in enumerate(row) if i < len(hdr)}
                for row in it]
    else:
        recs = list(csv.DictReader(open(src, encoding='utf-8-sig')))
        hdr = list(recs[0].keys()) if recs else []
    for col in (args.page_key_column, args.page_allow_column):
        if col not in hdr:
            sys.exit('--page-allow-from %s has no %r column (has: %s)'
                     % (src, col, ', '.join(hdr)))
    for rec in recs:
        pg = (rec.get(args.page_key_column) or '').strip()
        if not pg:
            continue
        seen_pages.add(pg)
        if (rec.get(args.page_allow_column) or '').strip() == args.page_allow_value:
            page_allow.add(pg)
    print('page allow-list: %d of %d page(s) labelled %r in %s'
          % (len(page_allow), len(seen_pages), args.page_allow_value, os.path.basename(src)))


def scope_reasons(source_name, page_slug):
    """Why this component is NOT in scope. Empty list means it is.

    A page the allow-list has never heard of is out of scope, not in it: the
    allow-list is the whole statement of what was assessed, so silence means
    'not assessed', and defaulting those IN would quietly re-admit the pages
    the narrowing was meant to exclude.
    """
    why = []
    if source_allow and source_name not in source_allow:
        why.append('source %s is not in the in-scope source list' % (source_name or '(none)'))
    if page_allow is not None and page_slug not in page_allow:
        why.append('page is not labelled %r in the page allow-list' % args.page_allow_value)
    return why


def in_scope(kind):
    return any(str(kind).startswith(k) for k in IN_SCOPE)


def page_url(pattern_id, slug):
    sub = subdomains.get(str(pattern_id))
    if not sub:
        return ''
    return 'https://%s.devmny.org/%s' % (sub, slug.lstrip('/'))


scans, comps = [], []
for f in args.scan:
    d = json.load(open(f, encoding='utf-8'))
    scans.append({
        'file': os.path.basename(f), 'patternId': d['patternId'],
        'scannedAt': d.get('scannedAt'), 'host': d.get('host'), 'app': d.get('app'),
        'pageCount': d.get('pageCount'), 'pagesScanned': d.get('pagesScanned'),
        'skippedNoSource': d.get('skippedNoSource'),
        'elementTypeCensus': d.get('elementTypeCensus') or {},
    })
    comps.extend(d['components'])

# stable, work-order sort: pattern, page, then position on the page
comps.sort(key=lambda c: (str(c['patternId']), c['pageSlug'], c['draftIndex']))

# ── Fix ID assignment ──────────────────────────────────────────────────────
# Sort order is for READABILITY (work order, by page). Identity must not depend
# on it: other staff add and remove components between scans, and on 2026-08-28
# a single rescan would have renumbered 193 of 497 rows - after the owner had
# already started working from those ids. The ledger pins Fix ID to the
# component itself, so ids are append-only and stable.
ledger, ledger_dirty = {}, False
if args.id_ledger and os.path.exists(args.id_ledger):
    ledger = json.load(open(args.id_ledger, encoding='utf-8'))

csv_path = os.path.join(args.out_dir, args.slug + '.csv')
prev = {}
if os.path.exists(csv_path):
    with open(csv_path, encoding='utf-8-sig') as f:
        for r in csv.DictReader(f):
            prev[(r.get('Fix ID'), r.get('Pattern ID'))] = r

# Reserve every id the LEDGER knows AND every id the existing report shows. The
# two sets differ: a component shared between two pages occupies two report rows
# but one ledger entry, so seeding the reserve from ledger values alone frees
# those ids and hands them to newly-added components. That happened on the first
# rebuild - 4 ids were re-minted. Reserve from both.
used_c, used_x = set(), set()
for v in list(ledger.values()) + [k[0] for k in prev]:
    if not v:
        continue
    if v.startswith('T6-C'):
        used_c.add(int(v[4:]))
    elif v.startswith('T6-X'):
        used_x.add(int(v[4:]))


def assign(pattern_id, section_id, prefix):
    """Reuse this component's existing Fix ID, or mint the next free one."""
    global ledger_dirty
    key = '%s:%s' % (pattern_id, section_id)
    if key in ledger:
        return ledger[key]
    pool = used_c if prefix == 'C' else used_x
    n = 1
    while n in pool:
        n += 1
    pool.add(n)
    fid = 'T6-%s%03d' % (prefix, n)
    ledger[key] = fid
    ledger_dirty = True
    return fid


rows, extras, deferred = [], [], []
n_in = n_out = 0
for c in comps:
    kind = c['elementType']
    cls = c['sourceClass']
    stored = c['storedFetchMode']
    resolved = c['resolvedFetchMode']
    want = WANT.get(cls)

    rec = {
        'Fix ID': '',
        'Parent finding': args.parent_finding,
        'Pattern ID': c['patternId'],
        'Page URL': page_url(c['patternId'], c['pageSlug']),
        'Page': c['pageSlug'],
        'Page ID': c['pageId'],
        'Section title': c['title'] if c['title'] else '(untitled %s)' % kind,
        'Draft section ID': c['sectionId'],
        'Tracking ID': c['trackingId'] or '',
        'Component kind': kind,
        'Draft index': c['draftIndex'],
        'Hidden from view': 'TRUE' if c['hideInView'] else '',
        'Source name': c['sourceName'] or '',
        'Source label (as shown in picker)': c['sourceLabel'],
        'Source class': cls,
        'Source ID': c['sourceId'] if c['sourceId'] is not None else '',
        'View ID': c['viewId'] if c['viewId'] is not None else '',
        'Source env (srcEnv)': c['srcEnv'] or '',
        'Has join': 'TRUE' if c['hasJoin'] else '',
        'Stored fetch mode': MODE_LABEL.get(stored, '(not set)') if stored else '(not set)',
        'readyToLoad': '' if c['readyToLoad'] is None else str(c['readyToLoad']).upper(),
        'Resolved behaviour': MODE_LABEL.get(resolved, resolved),
        'Relies on implicit fallback': 'TRUE' if c['reliesOnFallback'] else '',
        'In scope': '',
        'Scope reason': '',
        'Recommended fetch mode': '',
        'Recommended fix': '',
        'Fix needed': '',
        'Status': 'Open',
        'Assigned to': '',
        'Date fixed': '',
        'Notes': '',
    }

    if not in_scope(kind):
        n_out += 1
        rec['Fix ID'] = assign(c['patternId'], c['sectionId'], 'X')
        rec['In scope'] = 'No'
        rec['Scope reason'] = '%s exposes no Data Fetch Mode control' % kind
        rec['Recommended fetch mode'] = '(no control)'
        rec['Recommended fix'] = (
            'None available. %s exposes no Data Fetch Mode control (only Card, Spreadsheet and '
            'Graph do), so this component cannot be set by an author. It resolves through the '
            'implicit fallback and would need a library change to become settable.' % kind
        )
        rec['Fix needed'] = 'Out of scope - no setting exists'
        rec['Status'] = 'Out of scope'
        extras.append(rec)
        continue

    # ── scope gate ────────────────────────────────────────────────────────
    # Deferred rows keep their Fix ID and every scanned fact. `Force` costs a
    # query on every mount, so a component left alone here is a deliberate
    # decision with a reason attached, not an oversight.
    why = scope_reasons(c['sourceName'] or '', c['pageSlug'])
    if why:
        rec['Fix ID'] = assign(c['patternId'], c['sectionId'], 'C')
        rec['In scope'] = 'No'
        rec['Scope reason'] = '; '.join(why)
        rec['Recommended fetch mode'] = '(deferred)'
        rec['Recommended fix'] = (
            'None for now - deferred from the fetch-mode sweep because %s. Every component set to '
            '%s re-queries on every mount, so the sweep is deliberately limited. It keeps '
            'behaving as %s.' % ('; and '.join(why), MODE_LABEL['force'],
                                 MODE_LABEL.get(resolved, resolved))
        )
        if stored == 'force':
            # Written before the scope narrowed. Say so in the machine-owned column;
            # `Notes` belongs to the owner and is carried forward across rebuilds.
            rec['Recommended fix'] += (
                ' Already stored as %s - left as-is; reverting is not required.'
                % MODE_LABEL['force'])
        rec['Fix needed'] = 'Deferred - out of narrowed scope'
        rec['Status'] = 'Deferred'
        deferred.append(rec)
        continue

    n_in += 1
    rec['Fix ID'] = assign(c['patternId'], c['sectionId'], 'C')
    rec['In scope'] = 'Yes'
    rec['Recommended fetch mode'] = MODE_LABEL[want]
    if stored == want:
        rec['Recommended fix'] = 'None - already set to %s.' % MODE_LABEL[want]
        rec['Fix needed'] = 'No'
        rec['Status'] = 'Closed'
        rec['Date fixed'] = args.date
        rec['Notes'] = 'Already correct at scan time.'
    else:
        was = ('unset (behaving as %s via the implicit fallback, readyToLoad=%s)'
               % (MODE_LABEL.get(resolved, resolved), rec['readyToLoad'] or 'absent')) \
            if stored is None else 'set to %s' % MODE_LABEL.get(stored, stored)
        rec['Recommended fix'] = (
            'Section Settings > Data Fetch Mode: set %s (%s source). Currently %s.'
            % (MODE_LABEL[want], cls, was)
        )
        rec['Fix needed'] = 'Yes - set' if stored is None else 'Yes - change'
        if c['hideInView']:
            rec['Notes'] = ('Hidden from view (data.hideInView), so it never renders and the '
                            'setting has no runtime effect today - set it anyway for template '
                            'consistency, or defer. Owner decision.')
    rows.append(rec)

allrows = rows + deferred + extras
cols = list(allrows[0].keys())
os.makedirs(args.out_dir, exist_ok=True)

# ── rebuild safety ─────────────────────────────────────────────────────────
# Fix IDs are POSITIONAL (assigned by enumeration order), so adding or removing
# a component upstream would silently renumber every row after it - and the
# owner has already annotated rows by Fix ID. On a rebuild, check the existing
# CSV's Fix ID -> Draft section ID mapping and refuse if any id now points at a
# different component. Also carry forward the human-owned columns so a rebuild
# refreshes the machine-derived data without discarding triage.
CARRY = ('Assigned to', 'Notes')
if prev and not args.no_carry:
    remapped = []
    for r in allrows:
        old = prev.get((r['Fix ID'], r['Pattern ID']))
        if not old:
            continue
        if old.get('Draft section ID') != r['Draft section ID']:
            remapped.append((r['Fix ID'], old.get('Draft section ID'), r['Draft section ID']))
            continue
        for k in CARRY:
            if old.get(k) and not r[k]:
                r[k] = old[k]
    if remapped:
        for fid, was, now in remapped[:20]:
            print('REMAPPED %s: was section %s, now %s' % (fid, was, now))
        sys.exit('refusing to overwrite: %d Fix ID(s) now point at a different component. '
                 'Fix IDs are positional and the owner annotates by them - reconcile by hand, '
                 'or pass --no-carry to accept a renumbering.' % len(remapped))
    print('rebuild: %d prior row(s) matched, Fix IDs stable, %s carried forward'
          % (len(prev), '/'.join(CARRY)))

# A section shared between two pages produces two rows with ONE Fix ID - the
# ledger is keyed on the component, and that is the truthful reading: it is one
# component and one fix, listed under each page it appears on. Surface it rather
# than letting a duplicate id look like a bug. (county_template shares sections
# between lightning/wind and tornado/wind - see the T5 task.)
dupes = collections.defaultdict(list)
for r in allrows:
    dupes[(r['Fix ID'], r['Pattern ID'])].append(r['Page'])
shared = {k: v for k, v in dupes.items() if len(v) > 1}
if shared:
    print('note: %d Fix ID(s) appear on more than one page - one shared component, one fix:' % len(shared))
    for (fid, pid), pages in sorted(shared.items()):
        print('  %s (pattern %s) on %s' % (fid, pid, ', '.join(pages)))

if args.id_ledger and ledger_dirty:
    json.dump(ledger, open(args.id_ledger, 'w', encoding='utf-8'), indent=1, sort_keys=True)
    print('id ledger: %d assignment(s) -> %s' % (len(ledger), args.id_ledger))

# ── CSV ────────────────────────────────────────────────────────────────────
with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.DictWriter(f, fieldnames=cols)
    w.writeheader()
    w.writerows(allrows)

# ── XLSX ───────────────────────────────────────────────────────────────────
xlsx_path = os.path.join(args.out_dir, args.slug + '.xlsx')
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()

    def sheet(ws, data):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='37576B')
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        for r in data:
            ws.append([r[c] for c in cols])
        ws.freeze_panes = 'A2'
        widths = {'Recommended fix': 62, 'Source label (as shown in picker)': 42, 'Page': 46,
                  'Page URL': 46, 'Section title': 34, 'Notes': 46, 'Stored fetch mode': 24,
                  'Resolved behaviour': 24, 'Recommended fetch mode': 24, 'Source name': 30}
        for i, c in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 15)
        if data:
            ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(cols)), len(data) + 1)

    ws = wb.active
    ws.title = 'In scope - to fix'
    sheet(ws, rows)
    if deferred:
        sheet(wb.create_sheet('Deferred - out of scope'), deferred)
    sheet(wb.create_sheet('No fetch-mode control'), extras)
    wb.save(xlsx_path)
except ImportError:
    xlsx_path = None

# ── numbers for the HTML ───────────────────────────────────────────────────
def tally(data, *keys):
    return collections.Counter(tuple(r[k] for k in keys) for r in data)

by_pattern = collections.OrderedDict()
for r in rows:
    by_pattern.setdefault(r['Pattern ID'], []).append(r)

summary = []
for pid, rs in by_pattern.items():
    ext = [r for r in rs if r['Source class'] == 'external']
    itn = [r for r in rs if r['Source class'] == 'internal']
    summary.append({
        'pattern': pid,
        'subdomain': subdomains.get(str(pid), ''),
        'total': len(rs),
        'external': len(ext),
        'internal': len(itn),
        'writes': sum(1 for r in rs if r['Fix needed'].startswith('Yes')),
        'set': sum(1 for r in rs if r['Fix needed'] == 'Yes - set'),
        'change': sum(1 for r in rs if r['Fix needed'] == 'Yes - change'),
        'ok': sum(1 for r in rs if r['Fix needed'] == 'No'),
        'hidden': sum(1 for r in rs if r['Hidden from view'] == 'TRUE'),
        'kinds': collections.Counter(r['Component kind'] for r in rs),
        'deferred': sum(1 for r in deferred if r['Pattern ID'] == pid),
    })

E = html.escape


def badge(text, tone):
    tones = {
        'ext': 'bg-mny-100 text-mny-900 border-mny-200',
        'int': 'bg-mny-y50 text-mny-900 border-mny-y500',
        'ok': 'bg-emerald-50 text-emerald-900 border-emerald-200',
        'set': 'bg-amber-50 text-amber-900 border-amber-200',
        'chg': 'bg-rose-50 text-rose-900 border-rose-200',
        'off': 'bg-slate-100 text-slate-600 border-slate-300',
    }
    return ('<span class="inline-block whitespace-nowrap px-2 py-[1px] rounded border '
            'text-[11px] font-semibold tracking-wide %s">%s</span>' % (tones[tone], E(text)))


def mode_cell(label):
    if label.startswith('Force'):
        return badge('FORCE', 'int')
    if label.startswith('Smart'):
        return badge('SMART', 'ext')
    if label.startswith('Cache'):
        return badge('CACHE', 'chg')
    return badge('NOT SET', 'off')


def fix_cell(r):
    fn = r['Fix needed']
    if fn == 'No':
        return badge('already correct', 'ok')
    if fn == 'Yes - set':
        return badge('set', 'set')
    if fn == 'Yes - change':
        return badge('change', 'chg')
    return badge('no control', 'off')


parts = []
parts.append("""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>MitigateNY &mdash; County Template Data Fetch Mode (T6)</title>
<script src="https://cdn.tailwindcss.com"></script>
<script>
tailwind.config = { theme: { extend: {
  colors: {'mny-900':'#2D3E4C','mny-700':'#37576B','mny-400':'#6D96AE','mny-200':'#C5D7E0',
           'mny-100':'#E0EBF0','mny-50':'#F3F8F9','mny-y700':'#EAAD43','mny-y500':'#F1CA87',
           'mny-y50':'#FCF6EC','mny-red':'#DD524C','mny-redk':'#AA2E26','mny-grn':'#54B99B'},
  fontFamily: {display:['"Oswald"','sans-serif'], proxima:['"Source Sans 3"','system-ui','sans-serif']}
}}}
</script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Oswald:wght@400;500;600;700&family=Source+Sans+3:ital,wght@0,400;0,600;0,700;1,400&display=swap">
<style>
  body{font-family:'Source Sans 3',system-ui,sans-serif}
  h1,h2,h3,.display{font-family:'Oswald',sans-serif;letter-spacing:.01em}
  code{font-family:ui-monospace,'Cascadia Mono',Consolas,monospace;font-size:.86em;overflow-wrap:anywhere}
  .wrap{max-width:1500px}
  table{border-collapse:separate;border-spacing:0}
  thead th{position:sticky;top:0;z-index:2}
  tbody tr:hover{background:#F3F8F9}
  .num{font-variant-numeric:tabular-nums}
  @media print{.noprint{display:none}thead th{position:static}}
</style>
</head>
<body class="bg-[#F4F4F4] text-mny-900">
<div class="wrap mx-auto px-5 py-8">
""")

parts.append("""
<header class="mb-8 border-b-2 border-mny-700 pb-5">
  <div class="text-[11px] font-semibold tracking-[.18em] text-mny-400 uppercase">MitigateNY &middot; County template QA</div>
  <h1 class="text-4xl font-semibold mt-1">Data Fetch Mode &mdash; component catalog</h1>
  <p class="text-mny-700 mt-2 max-w-4xl">Every Card, Spreadsheet and Graph in the pattern, with the
  data source it binds, the <code>Data Fetch Mode</code> stored on it, the behaviour that actually
  runs, and the recommended fix. The per-component expansion of finding
  <strong>%s</strong> in <em>County Template QA</em>.</p>
</header>
""" % E(args.parent_finding))

# the rule + the scope it is applied within
scope_bits = []
if source_allow:
    scope_bits.append('<li>the data source is one of %s</li>'
                      % ', '.join('<code>%s</code>' % E(x) for x in sorted(source_allow)))
if page_allow is not None:
    scope_bits.append('<li>the page is one of the <strong>%d</strong> labelled <code>%s</code> in '
                      '<code>%s</code></li>'
                      % (len(page_allow), E(args.page_allow_value),
                         E(os.path.basename(args.page_allow_from))))

parts.append("""
<section class="mb-8 grid md:grid-cols-2 gap-4">
  <div class="bg-white border border-mny-200 rounded-lg p-5">
    <h2 class="text-lg font-semibold mb-2">The rule this report applies</h2>
    <table class="text-sm w-full">
      <tr class="border-b border-mny-100"><td class="py-1.5 pr-3">%s source</td>
        <td class="py-1.5">&rarr;&nbsp; %s</td></tr>
      <tr><td class="py-1.5 pr-3">%s source</td><td class="py-1.5">&rarr;&nbsp; %s</td></tr>
    </table>
    <p class="text-[13px] text-mny-700 mt-3">External sources are DAMA sources reached through a
    pgEnv (<code>hazmit_dama</code>); their content changes on a publication cycle, so
    <em>Smart</em> re-fetches when the query changes and otherwise reuses the cache. Internal
    sources are DMS-managed datasets in this app; an author can edit a row and expect to see it, so
    <em>Force</em> re-queries every mount.</p>
    <p class="text-[13px] text-mny-700 mt-2">With <code>fetchMode</code> absent the loader falls
    back to <code>readyToLoad === true ? 'smart' : 'cache'</code>
    (<code>dataWrapper/useDataLoader.js:245&ndash;247</code>), so an unset component already behaves
    as Smart wherever <code>readyToLoad</code> is true &mdash; invisibly, and only by accident of a
    second setting. <strong>Stored fetch mode</strong> is what an author sees in the panel;
    <strong>Resolved behaviour</strong> is what the browser does. A fix is judged against the stored
    value, because that is the thing that is missing.</p>
  </div>
  <div class="bg-white border %s rounded-lg p-5">
    <h2 class="text-lg font-semibold mb-2">%s</h2>
    %s
  </div>
</section>
""" % (badge('external', 'ext'), MODE_LABEL['smart'], badge('internal', 'int'), MODE_LABEL['force'],
       'border-mny-y500' if scope_bits else 'border-mny-200',
       'The scope this rule is applied within' if scope_bits else 'Scope: the whole pattern',
       ("""<p class="text-[13px] text-mny-700"><strong>Force re-queries on every mount</strong>, so
    each component set to it is page-load cost. The sweep is therefore limited to the datasets an
    author actually edits and expects to see change. A component is fixed only when:</p>
    <ul class="text-[13px] text-mny-700 mt-2 ml-4 list-disc space-y-1">%s</ul>
    <p class="text-[13px] text-mny-700 mt-3">The other <strong>%d</strong> components are
    <em>deferred</em>, not dropped: they keep their Fix ID and carry the reason, in the
    <code>Deferred - out of scope</code> sheet and the appendix below. Nothing about them is
    broken &mdash; they keep resolving through the implicit fallback exactly as they do today.</p>"""
        % (''.join(scope_bits), len(deferred)))
       if scope_bits else
       """<p class="text-[13px] text-mny-700">No source or page filter was applied: every data
    component in the pattern is in scope. Pass <code>--source-allow</code> and
    <code>--page-allow-from</code> to narrow the sweep.</p>"""))

# summary tiles
for s in summary:
    parts.append("""
<section class="mb-6">
  <h2 class="text-xl font-semibold mb-3">Pattern %s%s</h2>
  <div class="grid grid-cols-2 md:grid-cols-6 gap-3">
""" % (E(str(s['pattern'])), (' &middot; <code>%s</code>' % E(s['subdomain'])) if s['subdomain'] else ''))
    tiles = [
        ('components in scope', s['total'], 'text-mny-900'),
        ('writes needed', s['writes'], 'text-mny-redk'),
        ('already correct', s['ok'], 'text-emerald-700'),
        ('deferred out of scope', s['deferred'], 'text-slate-500'),
        ('hidden from view', s['hidden'], 'text-slate-500'),
        ('pages touched', len(set(r['Page'] for r in rs)), 'text-mny-700'),
    ]
    for label, val, tone in tiles:
        parts.append("""    <div class="bg-white border border-mny-200 rounded-lg px-4 py-3">
      <div class="text-3xl font-semibold num %s">%d</div>
      <div class="text-[11px] uppercase tracking-wider text-mny-400 mt-1">%s</div>
    </div>
""" % (tone, val, label))
    parts.append("""  </div>
  <p class="text-[13px] text-mny-700 mt-3">Of the %d writes: <strong>%d</strong> set a mode where
  none is stored, <strong>%d</strong> change one that is stored to a different value. By kind:
  %s.</p>
</section>
""" % (s['writes'], s['set'], s['change'],
       ', '.join('%d %s' % (v, k) for k, v in sorted(s['kinds'].items(), key=lambda kv: -kv[1]))))

# filter bar
parts.append("""
<section class="noprint bg-white border border-mny-200 rounded-lg p-4 mb-4 sticky top-0 z-10 shadow-sm">
  <div class="flex flex-wrap gap-3 items-end">
    <label class="flex flex-col"><span class="text-[11px] uppercase tracking-wider text-mny-400">Search page / title / source</span>
      <input id="q" type="search" placeholder="e.g. flooding, LHMP_IA, Overview"
             class="border border-mny-200 rounded px-2 py-1.5 text-sm w-72"></label>
    <label class="flex flex-col"><span class="text-[11px] uppercase tracking-wider text-mny-400">Source class</span>
      <select id="cls" class="border border-mny-200 rounded px-2 py-1.5 text-sm">
        <option value="">all</option><option value="external">external</option><option value="internal">internal</option></select></label>
    <label class="flex flex-col"><span class="text-[11px] uppercase tracking-wider text-mny-400">Kind</span>
      <select id="kind" class="border border-mny-200 rounded px-2 py-1.5 text-sm">
        <option value="">all</option><option>Card</option><option>Spreadsheet</option><option>Graph</option></select></label>
    <label class="flex flex-col"><span class="text-[11px] uppercase tracking-wider text-mny-400">Fix</span>
      <select id="fix" class="border border-mny-200 rounded px-2 py-1.5 text-sm">
        <option value="">all</option><option value="Yes - set">set</option>
        <option value="Yes - change">change</option><option value="No">already correct</option></select></label>
    <label class="flex items-center gap-1.5 text-sm pb-1.5">
      <input id="vis" type="checkbox" class="rounded border-mny-200"> hide the hidden-from-view rows</label>
    <div class="pb-1.5 text-sm text-mny-700"><span id="count" class="font-semibold num"></span> shown</div>
  </div>
</section>
""")

# main table
parts.append("""
<div class="bg-white border border-mny-200 rounded-lg overflow-x-auto mb-8">
<table class="w-full text-[13px]" id="tbl">
<thead class="bg-mny-700 text-white text-left">
  <tr>
    <th class="px-2.5 py-2 font-semibold">Fix ID</th>
    <th class="px-2.5 py-2 font-semibold" title="position in the page's draft_sections">Idx</th>
    <th class="px-2.5 py-2 font-semibold">Section title</th>
    <th class="px-2.5 py-2 font-semibold">Draft&nbsp;ID</th>
    <th class="px-2.5 py-2 font-semibold">Kind</th>
    <th class="px-2.5 py-2 font-semibold">Data source</th>
    <th class="px-2.5 py-2 font-semibold">Class</th>
    <th class="px-2.5 py-2 font-semibold">Stored</th>
    <th class="px-2.5 py-2 font-semibold">Resolved</th>
    <th class="px-2.5 py-2 font-semibold">Recommended</th>
    <th class="px-2.5 py-2 font-semibold">Fix</th>
  </tr>
</thead>
<tbody>
""")

last_page = None
for r in rows:
    if r['Page'] != last_page:
        last_page = r['Page']
        n = sum(1 for x in rows if x['Page'] == last_page)
        url = r['Page URL']
        link = ('<a class="underline decoration-mny-200 hover:decoration-mny-700" href="%s" '
                'target="_blank" rel="noreferrer">%s</a>' % (E(url), E(last_page))) if url else E(last_page)
        parts.append('<tr class="bg-mny-50 border-y border-mny-200"><td colspan="11" '
                     'class="px-2.5 py-1.5 font-semibold text-mny-700">%s '
                     '<span class="text-mny-400 font-normal num">&middot; %d component%s '
                     '&middot; page %s</span></td></tr>\n'
                     % (link, n, '' if n == 1 else 's', E(r['Page ID'])))
    hidden = r['Hidden from view'] == 'TRUE'
    parts.append(
        '<tr class="border-b border-mny-100 align-top" data-cls="%s" data-kind="%s" data-fix="%s" '
        'data-hidden="%s" data-q="%s">'
        '<td class="px-2.5 py-1.5 num text-mny-400 whitespace-nowrap">%s</td>'
        '<td class="px-2.5 py-1.5 num text-mny-400">%s</td>'
        '<td class="px-2.5 py-1.5">%s%s</td>'
        '<td class="px-2.5 py-1.5 num whitespace-nowrap"><code>%s</code></td>'
        '<td class="px-2.5 py-1.5 whitespace-nowrap">%s</td>'
        '<td class="px-2.5 py-1.5"><code>%s</code></td>'
        '<td class="px-2.5 py-1.5">%s</td>'
        '<td class="px-2.5 py-1.5">%s</td>'
        '<td class="px-2.5 py-1.5">%s</td>'
        '<td class="px-2.5 py-1.5">%s</td>'
        '<td class="px-2.5 py-1.5">%s</td>'
        '</tr>\n' % (
            E(r['Source class']), E(r['Component kind'].split(':')[0]), E(r['Fix needed']),
            'y' if hidden else 'n',
            E((r['Page'] + ' ' + r['Section title'] + ' ' + r['Source name'] + ' ' + r['Draft section ID']).lower()),
            E(r['Fix ID']), E(str(r['Draft index'])),
            E(r['Section title']),
            ' <span class="text-[10px] uppercase tracking-wider text-slate-400">hidden</span>' if hidden else '',
            E(r['Draft section ID']), E(r['Component kind']),
            E(r['Source label (as shown in picker)']),
            badge(r['Source class'], 'ext' if r['Source class'] == 'external' else 'int'),
            mode_cell(r['Stored fetch mode']), mode_cell(r['Resolved behaviour']),
            mode_cell(r['Recommended fetch mode']), fix_cell(r),
        ))
parts.append('</tbody></table></div>\n')

# deferred appendix - grouped by reason, then by page, so the shape of what was
# set aside is readable without scrolling 400 rows
if deferred:
    by_reason = collections.OrderedDict()
    for r in deferred:
        by_reason.setdefault(r['Scope reason'], []).append(r)
    parts.append("""
<section class="mb-8">
  <h2 class="text-xl font-semibold mb-2">Appendix &mdash; %d components deferred from the narrowed scope</h2>
  <p class="text-[13px] text-mny-700 mb-3 max-w-4xl">Listed rather than dropped, because
  &ldquo;we decided not to set this&rdquo; and &ldquo;we never looked at it&rdquo; must not read the
  same. Each keeps its Fix ID, so widening the scope later is a rebuild and not a renumbering. None
  of these is broken: with no <code>fetchMode</code> stored they resolve through the implicit
  fallback, which is what they already do today.</p>
  <div class="bg-white border border-mny-200 rounded-lg overflow-x-auto">
  <table class="w-full text-[13px]"><thead class="bg-slate-500 text-white text-left"><tr>
    <th class="px-2.5 py-2 font-semibold">Deferred because</th>
    <th class="px-2.5 py-2 font-semibold num">Components</th>
    <th class="px-2.5 py-2 font-semibold">Sources</th>
    <th class="px-2.5 py-2 font-semibold">Pages</th>
    <th class="px-2.5 py-2 font-semibold">Stored today</th></tr></thead><tbody>
""" % len(deferred))
    for reason, rs in sorted(by_reason.items(), key=lambda kv: -len(kv[1])):
        srcs = collections.Counter(r['Source name'] or '(none)' for r in rs)
        pgs = collections.Counter(r['Page'] for r in rs)
        stored_c = collections.Counter(r['Stored fetch mode'] for r in rs)
        def top(c, n=4):
            head = ', '.join('%s <span class="text-mny-400 num">%d</span>' % (E(k), v)
                             for k, v in c.most_common(n))
            rest = len(c) - n
            return head + ('<span class="text-mny-400"> &hellip; +%d more</span>' % rest
                           if rest > 0 else '')
        parts.append('<tr class="border-b border-mny-100 align-top">'
                     '<td class="px-2.5 py-1.5">%s</td>'
                     '<td class="px-2.5 py-1.5 num font-semibold">%d</td>'
                     '<td class="px-2.5 py-1.5"><code>%s</code></td>'
                     '<td class="px-2.5 py-1.5">%s</td>'
                     '<td class="px-2.5 py-1.5">%s</td></tr>\n'
                     % (E(reason), len(rs), top(srcs), top(pgs), top(stored_c)))
    parts.append('</tbody></table></div></section>\n')

# out-of-scope appendix
if extras:
    parts.append("""
<section class="mb-8">
  <h2 class="text-xl font-semibold mb-2">Appendix &mdash; %d components that bind a source but have no Data Fetch Mode control</h2>
  <p class="text-[13px] text-mny-700 mb-3 max-w-4xl">These are listed so the catalog's boundary is
  visible rather than implicit. Only <code>Card.config.jsx</code>,
  <code>spreadsheet/config.jsx</code>, <code>graph/config.jsx</code> and
  <code>graph_new/config.jsx</code> register a <code>Data Fetch Mode</code> select;
  <code>FilterComponent.config.js</code> and <code>header.config.js</code> do not. So these
  components resolve purely through the implicit fallback and <strong>cannot be fixed by an
  author</strong> &mdash; making them settable is a library change, not a content one.</p>
  <div class="bg-white border border-mny-200 rounded-lg overflow-x-auto">
  <table class="w-full text-[13px]"><thead class="bg-slate-500 text-white text-left"><tr>
    <th class="px-2.5 py-2 font-semibold">Fix ID</th><th class="px-2.5 py-2 font-semibold">Page</th>
    <th class="px-2.5 py-2 font-semibold">Draft ID</th><th class="px-2.5 py-2 font-semibold">Kind</th>
    <th class="px-2.5 py-2 font-semibold">Data source</th><th class="px-2.5 py-2 font-semibold">Class</th>
    <th class="px-2.5 py-2 font-semibold">Resolved behaviour</th></tr></thead><tbody>
""" % len(extras))
    for r in extras:
        parts.append('<tr class="border-b border-mny-100"><td class="px-2.5 py-1.5 num text-mny-400">%s</td>'
                     '<td class="px-2.5 py-1.5"><code>%s</code></td>'
                     '<td class="px-2.5 py-1.5 num"><code>%s</code></td>'
                     '<td class="px-2.5 py-1.5">%s</td><td class="px-2.5 py-1.5"><code>%s</code></td>'
                     '<td class="px-2.5 py-1.5">%s</td><td class="px-2.5 py-1.5">%s</td></tr>\n'
                     % (E(r['Fix ID']), E(r['Page']), E(r['Draft section ID']), E(r['Component kind']),
                        E(r['Source label (as shown in picker)']),
                        badge(r['Source class'], 'ext' if r['Source class'] == 'external' else 'int'),
                        mode_cell(r['Resolved behaviour'])))
    parts.append('</tbody></table></div></section>\n')

# provenance
parts.append('<section class="mb-8 bg-white border border-mny-200 rounded-lg p-5">'
             '<h2 class="text-lg font-semibold mb-2">How this was built</h2>'
             '<table class="text-[13px] w-full"><tbody>')
for s in scans:
    parts.append('<tr class="border-b border-mny-100"><td class="py-1.5 pr-4 whitespace-nowrap">'
                 '<code>%s</code></td><td class="py-1.5">pattern <strong>%s</strong> &middot; '
                 '%s of %s pages &middot; scanned %s &middot; %s sections carried no '
                 '<code>externalSource</code> and are not data components</td></tr>\n'
                 % (E(s['file']), E(str(s['patternId'])), E(str(s['pagesScanned'])),
                    E(str(s['pageCount'])), E(str(s['scannedAt'])), E(str(s['skippedNoSource']))))
parts.append('</tbody></table>'
             '<p class="text-[13px] text-mny-700 mt-3">Scanned with '
             '<code>planning/mitigateny/skills/scripts/report_fixes/scan_fetchmode.mjs</code> over '
             'every page of the pattern\'s <code>draft_sections</code>, and rendered by '
             '<code>build_fetchmode_report.py</code> in the same folder. Draft sections only &mdash; '
             'never published ids; see the report-fix skill for why.</p></section>')

parts.append("""
<script>
(function(){
  const q=document.getElementById('q'), cls=document.getElementById('cls'),
        kind=document.getElementById('kind'), fix=document.getElementById('fix'),
        vis=document.getElementById('vis'), count=document.getElementById('count'),
        rows=[...document.querySelectorAll('#tbl tbody tr[data-q]')],
        heads=[...document.querySelectorAll('#tbl tbody tr:not([data-q])')];
  function apply(){
    const t=q.value.trim().toLowerCase();
    let n=0;
    for(const r of rows){
      const ok = (!t || r.dataset.q.includes(t))
        && (!cls.value || r.dataset.cls===cls.value)
        && (!kind.value || r.dataset.kind===kind.value)
        && (!fix.value || r.dataset.fix===fix.value)
        && (!vis.checked || r.dataset.hidden==='n');
      r.style.display = ok ? '' : 'none';
      if(ok) n++;
    }
    // a page heading disappears when nothing under it survives the filter
    for(const h of heads){
      let any=false;
      for(let s=h.nextElementSibling; s && s.hasAttribute('data-q'); s=s.nextElementSibling)
        if(s.style.display!=='none'){any=true;break;}
      h.style.display = any ? '' : 'none';
    }
    count.textContent = n;
  }
  [q,cls,kind,fix,vis].forEach(el=>el.addEventListener('input',apply));
  apply();
})();
</script>
</div></body></html>
""")

html_path = os.path.join(args.out_dir, args.slug + '.html')
open(html_path, 'w', encoding='utf-8').write(''.join(parts))

print('rows in scope      : %d  on %d page(s)'
      % (len(rows), len(set(r['Page'] for r in rows))))
print('rows deferred      : %d  (narrowed scope)' % len(deferred))
print('rows with no control: %d' % len(extras))
print('writes recommended : %d (set %d, change %d)'
      % (sum(1 for r in rows if r['Fix needed'].startswith('Yes')),
         sum(1 for r in rows if r['Fix needed'] == 'Yes - set'),
         sum(1 for r in rows if r['Fix needed'] == 'Yes - change')))
print('already correct    : %d' % sum(1 for r in rows if r['Fix needed'] == 'No'))
print('csv  -> %s' % csv_path)
print('xlsx -> %s' % (xlsx_path or '(openpyxl not installed - skipped)'))
print('html -> %s' % html_path)
