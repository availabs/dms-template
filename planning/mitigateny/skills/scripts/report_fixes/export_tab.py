# -*- coding: utf-8 -*-
"""
STEP 0 of the report-fix loop: freeze the exact report rows you are about to
act on.

Reports in src/themes/mny/design/reports/ are multi-tab workbooks; a fix run
should be driven by ONE tab, copied into the run folder so the run is
reproducible even after the report is regenerated.

usage:
  python export_tab.py <report.xlsx> "<Worksheet tab>" <out.csv>
  python export_tab.py --list <report.xlsx>
"""
import sys, csv
import openpyxl

if sys.argv[1] == '--list':
    wb = openpyxl.load_workbook(sys.argv[2], read_only=True)
    for ws in wb.worksheets:
        print('%-32s rows %5d  cols %d' % (ws.title, ws.max_row - 1, ws.max_column))
    sys.exit(0)

src, tab, out = sys.argv[1], sys.argv[2], sys.argv[3]
wb = openpyxl.load_workbook(src, read_only=True)
if tab not in wb.sheetnames:
    sys.exit('no such tab %r; have: %s' % (tab, ', '.join(wb.sheetnames)))

rows = list(wb[tab].iter_rows(values_only=True))
hdr = [('' if h is None else str(h)) for h in rows[0]]
body = [r for r in rows[1:] if r[0] not in (None, '')]

with open(out, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.writer(f)
    w.writerow(hdr)
    for r in body:
        w.writerow(['' if v is None else v for v in r])

print('%s :: %s -> %s (%d rows)' % (src.split('/')[-1], tab, out, len(body)))
